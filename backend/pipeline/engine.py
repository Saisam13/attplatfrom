"""Chemical Trading Attractiveness pipeline — ported from trading_module.py v2.0.

Stages: 1 ingest -> 2 NLP match (+optional LLM assist) -> 3 aggregate ->
4 analyse -> 4b geo adjust -> 5 regulatory -> 6 composite score.

Produces the exact same numbers as the reference implementation for the same
inputs when the LLM assist is off (default).
"""
from . import patches  # noqa: F401  (openpyxl aRGB monkey-patch — must be first)

import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from statistics import median, stdev, mean

import openpyxl

from .constants import (
    WEIGHTS, NLP_DIRECT_THRESHOLD, NLP_NEAR_THRESHOLD, TIER_A_MIN, TIER_B_MIN,
    OUTLIER_IQR_MULT, WINSORIZE_LO, WINSORIZE_HI, GEO_EVENTS, REGULATORY,
    NOISE_RE, UNIT_TO_KG, SYNONYMS, SPELLING, EASE,
)


# ══════════════════════════════════════════════════════════════
# Text / unit helpers
# ══════════════════════════════════════════════════════════════
def _clean_description(desc):
    d = NOISE_RE.sub(' ', desc)
    d = re.sub(r'\d{2,}\s*KGS?\b', ' ', d)
    d = re.sub(r'\d+\s*(?:DRUMS?|BAGS?|PALLETS?|PACKAGES?|CARTONS?)\b', ' ', d)
    d = re.sub(r'\bMATERIAL\s+NO\.?\s*\d+', ' ', d)
    d = re.sub(r'\s+', ' ', d).strip()
    return d


def _normalize_to_kg(qty, unit, desc):
    if qty <= 0:
        return 0
    u = unit.upper().strip()
    if u in UNIT_TO_KG:
        return qty * UNIT_TO_KG[u]
    if 'DRUM' in u:
        return qty * 200
    if 'BAG' in u:
        return qty * 25
    if 'PALLET' in u:
        return qty * 1000
    if 'BOTTLE' in u or 'JAR' in u:
        return qty * 1
    if 'PIECE' in u or 'PCS' in u or 'EA' in u or 'UNIT' in u:
        m = re.search(r'(\d+(?:\.\d+)?)\s*KG', desc)
        if m:
            return qty * float(m.group(1))
        return qty * 1
    return qty


# ══════════════════════════════════════════════════════════════
# Outlier handling
# ══════════════════════════════════════════════════════════════
def _trim_outliers(prices):
    if len(prices) < 4:
        return prices[:], 0
    s = sorted(prices)
    n = len(s)
    q1 = s[n // 4]
    q3 = s[3 * n // 4]
    iqr = q3 - q1
    lo = q1 - OUTLIER_IQR_MULT * iqr
    hi = q3 + OUTLIER_IQR_MULT * iqr
    clean = [p for p in prices if lo <= p <= hi]
    return clean, len(prices) - len(clean)


def _winsorize(prices):
    if len(prices) < 10:
        return prices[:]
    s = sorted(prices)
    n = len(s)
    lo_idx = max(0, int(n * WINSORIZE_LO / 100))
    hi_idx = min(n - 1, int(n * WINSORIZE_HI / 100))
    lo_val = s[lo_idx]
    hi_val = s[hi_idx]
    return [max(lo_val, min(hi_val, p)) for p in prices]


def _clean_prices(prices):
    trimmed, n_removed = _trim_outliers(prices)
    return _winsorize(trimmed), n_removed


def _iqr(values):
    if len(values) < 2:
        return 0
    s = sorted(values)
    n = len(s)
    return s[3 * n // 4] - s[n // 4]


# ══════════════════════════════════════════════════════════════
# STAGE 1 — DATA INGESTION
# ══════════════════════════════════════════════════════════════
def load_base_portfolio(base_file):
    """Read base chemical portfolio from the 'Category Overview' sheet."""
    base_chemicals = []
    wb_base = openpyxl.load_workbook(base_file, read_only=True, data_only=True)
    sheet = 'Category Overview' if 'Category Overview' in wb_base.sheetnames else wb_base.sheetnames[0]
    ws = wb_base[sheet]
    section = None
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        v = [str(x).strip() if x else '' for x in row]
        if len(v) < 1:
            continue
        txt = v[0].upper()
        if 'METALLURG' in txt:
            section = 'MC'
            continue
        if 'BEAUTY' in txt and 'PERSONAL' in txt:
            section = 'BPC'
            continue
        if section == 'MC' and len(v) >= 4 and v[0].replace('.0', '').isdigit() and v[1]:
            hsn = v[3].replace('.0', '') if v[3] else ''
            base_chemicals.append({
                'name': v[1], 'cas': v[2] if len(v) > 2 else '', 'hsn': hsn,
                'chem_type': v[4] if len(v) > 4 else '',
                'name_lower': v[1].lower().strip(),
                'name_tokens': set(re.split(r'[\s\-\(\),/]+', v[1].lower())),
            })
    wb_base.close()
    return base_chemicals


def parse_exim_files(exim_files, log=print):
    """Parse EXIM xlsx files (row 1 title, row 3 headers, data from row 4)
    into normalized row dicts. Returns (rows, skipped_files)."""
    exim_rows = []
    skipped_files = []
    for fpath in sorted(exim_files):
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except PermissionError:
            log(f"    SKIPPED (file locked): {fname}")
            skipped_files.append(fname)
            continue
        except Exception as e:
            log(f"    SKIPPED (unreadable: {e}): {fname}")
            skipped_files.append(fname)
            continue
        ws = wb[wb.sheetnames[0]]
        file_rows = 0
        hsn6 = '?'
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            v = list(row)
            if len(v) < 13:
                continue
            date_val = v[0]
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m')
                date_obj = date_val
            elif date_val:
                try:
                    date_obj = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')
                    date_str = date_obj.strftime('%Y-%m')
                except Exception:
                    continue
            else:
                continue
            hsn_raw = str(v[1]).replace('.0', '').strip() if v[1] else ''
            if not hsn_raw or len(hsn_raw) < 4:
                continue
            hsn6 = hsn_raw[:6]
            desc = str(v[2]).strip().upper() if v[2] else ''
            if not desc:
                continue
            seller = str(v[3]).strip() if v[3] else 'N/A'
            seller_country = str(v[4]).strip().upper() if v[4] else ''
            buyer = str(v[6]).strip() if v[6] else 'N/A'
            buyer_country = str(v[7]).strip().upper() if v[7] else ''
            unit = str(v[9]).strip().upper() if v[9] else ''
            qty = 0
            try:
                q = v[10]
                if q and str(q).strip() not in ('', '-', 'N/A', 'None'):
                    qty = float(str(q).replace(',', ''))
            except Exception:
                pass
            value_usd = 0
            try:
                val = v[11]
                if val and str(val).strip() not in ('', '-', 'N/A', 'None'):
                    value_usd = float(str(val).replace(',', ''))
            except Exception:
                pass
            unit_price = 0
            try:
                up = v[12]
                if up and str(up).strip() not in ('', '-', 'N/A', 'None'):
                    unit_price = float(str(up).replace(',', ''))
            except Exception:
                pass
            qty_kg = _normalize_to_kg(qty, unit, desc)
            desc_clean = _clean_description(desc)
            rodtep = str(v[13]).strip() if len(v) > 13 and v[13] else ''
            drawback_rate = str(v[15]).strip() if len(v) > 15 and v[15] else ''
            exim_rows.append({
                'date': date_str, 'date_obj': date_obj, 'hsn6': hsn6, 'hsn_raw': hsn_raw,
                'desc_raw': desc, 'desc_clean': desc_clean,
                'seller': seller, 'seller_country': seller_country,
                'origin_port': str(v[5]).strip() if v[5] else '',
                'buyer': buyer, 'buyer_country': buyer_country,
                'dest_port': str(v[8]).strip() if v[8] else '',
                'unit': unit, 'qty': qty, 'qty_kg': qty_kg,
                'value_usd': value_usd, 'unit_price': unit_price,
                'rodtep': rodtep, 'drawback_rate': drawback_rate, 'file': fname,
            })
            file_rows += 1
        wb.close()
        log(f"    {fname[-25:]}: {file_rows} rows, HSN={hsn6}")
    return exim_rows, skipped_files


def stage1_ingest(exim_files, base_file, log=print):
    log("STAGE 1 — Ingesting data...")
    base_chemicals = load_base_portfolio(base_file)
    log(f"  Base chemicals loaded: {len(base_chemicals)}")

    hsn_to_base = defaultdict(list)
    for bc in base_chemicals:
        if bc['hsn']:
            hsn_to_base[bc['hsn'][:6]].append(bc)

    log(f"  EXIM files: {len(exim_files)}")
    exim_rows, skipped_files = parse_exim_files(exim_files, log)
    log(f"  Total EXIM rows: {len(exim_rows)}")
    return exim_rows, base_chemicals, hsn_to_base, skipped_files


# ══════════════════════════════════════════════════════════════
# STAGE 2 — NLP MATCHING (+ optional LLM assist)
# ══════════════════════════════════════════════════════════════
def _match_score(desc, name_lower, name_tokens):
    if name_lower in desc:
        return 0.95
    for syn, canonical in SYNONYMS.items():
        if syn in desc and canonical == name_lower:
            return 0.90
    desc_tokens = set(re.split(r'[\s\-\(\),/\.]+', desc)) - {'', 'of', 'the', 'and', 'with', 'in', 'for', 'or', 'a'}
    if not name_tokens:
        return 0
    overlap = name_tokens & desc_tokens
    if not overlap:
        desc_n = desc
        name_n = ' '.join(name_tokens)
        for old, new in SPELLING.items():
            desc_n = desc_n.replace(old, new)
            name_n = name_n.replace(old, new)
        overlap = set(name_n.split()) & set(re.split(r'[\s\-\(\),/\.]+', desc_n))
    token_ratio = len(overlap) / len(name_tokens)
    seq_score = SequenceMatcher(None, name_lower, desc[:len(name_lower) * 3]).ratio()
    return 0.6 * token_ratio + 0.4 * seq_score


def _extract_chem_name(desc):
    d = re.sub(r'\b(?:TOTAL|CONTAINING|MATERIAL|PRODUCT|CHEMICAL|COMPOUND|SOLUTION|MIXTURE|GRADE|PURE|REAGENT|TECHNICAL|INDUSTRIAL|POWDER|LIQUID|SOLID|CRYSTAL|GRANULAR|ANHYDROUS|HYDRATE|HEXAHYDRATE|MONOHYDRATE)\b', '', desc, flags=re.I)
    d = re.sub(r'\b\d+\.?\d*\s*%?\b', '', d)
    d = re.sub(r'\s+', ' ', d).strip()
    words = d.split()[:6]
    name = ' '.join(words)
    return name.title() if name else 'Unknown'


def stage2_nlp_match(exim_rows, base_chemicals, hsn_to_base, log=print, llm_matcher=None):
    """Rule-based fuzzy matching. If llm_matcher is provided, descriptions that
    score below the direct threshold are sent (batched, cached) for LLM
    identification against the base portfolio."""
    log("\nSTAGE 2 — NLP chemical matching...")
    base_lookup = []
    for bc in base_chemicals:
        tokens = bc['name_tokens'] - {'', 'of', 'the', 'and', 'with', 'in', 'for', 'or', 'a'}
        base_lookup.append({
            'name': bc['name'], 'name_lower': bc['name_lower'], 'tokens': tokens,
            'hsn6': bc['hsn'][:6] if bc['hsn'] else '', 'cas': bc['cas'],
        })
    match_stats = Counter()
    for row in exim_rows:
        desc = row['desc_clean'].lower()
        hsn6 = row['hsn6']
        candidates = [b for b in base_lookup if b['hsn6'] == hsn6]
        if not candidates:
            candidates = base_lookup
        best_score = 0
        best_match = None
        for bc in candidates:
            score = _match_score(desc, bc['name_lower'], bc['tokens'])
            if score > best_score:
                best_score = score
                best_match = bc
        if best_score >= NLP_DIRECT_THRESHOLD:
            row['chemical_id'] = best_match['name']
            row['match_type'] = 'direct'
        elif best_score >= NLP_NEAR_THRESHOLD:
            row['chemical_id'] = best_match['name']
            row['match_type'] = 'near'
        else:
            row['chemical_id'] = _extract_chem_name(desc)
            row['match_type'] = 'none'
        row['match_score'] = best_score
        match_stats[row['match_type']] += 1

    # HYBRID mode: LLM assist for sub-direct-threshold descriptions
    if llm_matcher is not None:
        sub = [r for r in exim_rows if r['match_score'] < NLP_DIRECT_THRESHOLD]
        uniq = sorted({r['desc_clean'] for r in sub if r['desc_clean']})
        if uniq:
            base_names = [b['name'] for b in base_lookup]
            mapping = llm_matcher(uniq, base_names)  # dict desc_clean -> base name | None
            valid = {b['name'] for b in base_lookup}
            n_llm = 0
            for r in sub:
                hit = mapping.get(r['desc_clean'])
                if hit and hit in valid:
                    match_stats[r['match_type']] -= 1
                    r['chemical_id'] = hit
                    r['match_type'] = 'llm'
                    r['match_score'] = max(r['match_score'], NLP_DIRECT_THRESHOLD)
                    n_llm += 1
            match_stats['llm'] = n_llm
            log(f"  LLM assist: {n_llm} rows re-matched ({len(uniq)} unique descriptions considered)")

    total = max(len(exim_rows), 1)
    log(f"  Direct: {match_stats['direct']} ({match_stats['direct']*100//total}%), "
        f"Near: {match_stats['near']} ({match_stats['near']*100//total}%), "
        f"None: {match_stats['none']} ({match_stats['none']*100//total}%)")
    return exim_rows, dict(match_stats)


# ══════════════════════════════════════════════════════════════
# STAGE 3 — AGGREGATION
# ══════════════════════════════════════════════════════════════
def _make_chem_bucket():
    return {
        'rows': [], 'total_qty_kg': 0, 'total_value_usd': 0, 'shipment_count': 0,
        'buyers': Counter(), 'buyer_countries': Counter(),
        'sellers': Counter(), 'seller_countries': Counter(),
        'monthly_qty': Counter(), 'monthly_value': Counter(), 'monthly_shipments': Counter(),
        'unit_prices': [], 'unit_prices_clean': [], 'outliers_removed': 0,
        'price_by_country': defaultdict(list), 'price_by_month': defaultdict(list),
        'hsn_codes': set(), 'match_types': Counter(), 'in_base': False,
        'rodtep_values': [], 'drawback_values': [],
    }


def _aggregate_rows(rows, base_names):
    chemicals = defaultdict(_make_chem_bucket)
    for row in rows:
        cid = row['chemical_id']
        c = chemicals[cid]
        c['rows'].append(row)
        c['total_qty_kg'] += row['qty_kg']
        c['total_value_usd'] += row['value_usd']
        c['shipment_count'] += 1
        if row['buyer'] and row['buyer'] != 'N/A':
            c['buyers'][row['buyer']] += 1
        if row['buyer_country']:
            c['buyer_countries'][row['buyer_country']] += 1
        if row['seller'] and row['seller'] != 'N/A':
            c['sellers'][row['seller']] += 1
        if row['seller_country']:
            c['seller_countries'][row['seller_country']] += 1
        c['monthly_qty'][row['date']] += row['qty_kg']
        c['monthly_value'][row['date']] += row['value_usd']
        c['monthly_shipments'][row['date']] += 1
        if row['unit_price'] > 0:
            c['unit_prices'].append(row['unit_price'])
            if row['seller_country']:
                c['price_by_country'][row['seller_country']].append(row['unit_price'])
            c['price_by_month'][row['date']].append(row['unit_price'])
        c['hsn_codes'].add(row['hsn6'])
        c['match_types'][row['match_type']] += 1
        c['in_base'] = cid.lower() in base_names
        if row['rodtep'] and row['rodtep'] not in ('-', ' - ', ''):
            try:
                c['rodtep_values'].append(float(row['rodtep']))
            except Exception:
                pass
        if row['drawback_rate'] and row['drawback_rate'] not in ('-', ' - ', ''):
            try:
                c['drawback_values'].append(float(row['drawback_rate']))
            except Exception:
                pass
    for cid, c in chemicals.items():
        c['unit_prices_clean'], c['outliers_removed'] = _clean_prices(c['unit_prices'])
    return dict(chemicals)


def stage3_aggregate(exim_rows, base_chemicals, log=print):
    log("\nSTAGE 3 — Aggregating by chemical...")
    base_names = {bc['name'].lower() for bc in base_chemicals}
    matched_rows = [r for r in exim_rows if r['match_type'] in ('direct', 'near', 'llm')]
    unmatched_rows = [r for r in exim_rows if r['match_type'] == 'none']

    base_chemicals_agg = _aggregate_rows(matched_rows, base_names)
    opp_chemicals_raw = _aggregate_rows(unmatched_rows, base_names)
    opp_chemicals = {k: v for k, v in opp_chemicals_raw.items() if v['shipment_count'] >= 3}

    log(f"  Base-matched chemicals: {len(base_chemicals_agg)}")
    log(f"  Opportunity chemicals (>=3 shipments): {len(opp_chemicals)}")
    return base_chemicals_agg, opp_chemicals


# ══════════════════════════════════════════════════════════════
# STAGE 4 — EIGHT ANALYSES
# ══════════════════════════════════════════════════════════════
def _score_volume(c):
    return c['total_qty_kg']


def _score_price(c):
    prices = c.get('unit_prices_clean', c.get('unit_prices', []))
    if not prices:
        return 0
    return median(prices) * c['total_qty_kg']


def _score_buyers(c):
    n_buyers = len(c['buyers'])
    n_countries = len(c['buyer_countries'])
    total = c['shipment_count']
    repeat_pct = sum(1 for cnt in c['buyers'].values() if cnt > 1) / max(n_buyers, 1)
    if total > 0:
        hhi = sum((cnt / total) ** 2 for cnt in c['buyers'].values())
        frag = 1 - hhi
    else:
        frag = 0
    return 0.3 * n_buyers + 0.3 * frag * 100 + 0.2 * n_countries * 5 + 0.2 * repeat_pct * 100


def _score_suppliers(c):
    n_sellers = len(c['sellers'])
    n_countries = len(c['seller_countries'])
    india_pct = c['seller_countries'].get('INDIA', 0) / max(c['shipment_count'], 1)
    return 0.4 * n_sellers + 0.3 * india_pct * 100 + 0.3 * n_countries * 5


def _score_trend(c):
    months = sorted(m for m in c['monthly_qty'].keys() if c['monthly_qty'][m] > 0)
    if len(months) < 3:
        return 50
    months = months[-12:]
    vals = [c['monthly_qty'][m] for m in months]
    n = len(vals)
    x = list(range(n))
    x_mean = mean(x)
    y_mean = mean(vals)
    num = sum((x[i] - x_mean) * (vals[i] - y_mean) for i in range(n))
    den = sum((x[i] - x_mean) ** 2 for i in range(n))
    if den == 0 or y_mean == 0:
        return 50
    growth_rate = (num / den) / y_mean
    return max(0, min(100, 50 + growth_rate * 500))


def _score_structure(c):
    total = c['shipment_count']
    n_sellers = len(c['sellers'])
    if total > 0 and n_sellers > 0:
        hhi = sum((cnt / total) ** 2 for cnt in c['sellers'].values())
        frag = 1 - hhi
    else:
        frag = 0
    months = sorted(m for m in c['monthly_shipments'].keys() if c['monthly_shipments'][m] > 0)
    freq = total / max(len(months), 1)
    commoditization = min(1, freq / 50)
    return 0.5 * frag * 100 + 0.5 * commoditization * 100


def _score_freedom(c):
    if not c['seller_countries'] and not c['buyer_countries']:
        return 50
    total = 0
    weight = 0
    for country, cnt in c['seller_countries'].items():
        total += EASE.get(country, 60) * cnt
        weight += cnt
    for country, cnt in c['buyer_countries'].items():
        total += EASE.get(country, 60) * cnt
        weight += cnt
    return total / max(weight, 1)


def _score_barrier(c):
    penalty = 0
    hazmat_count = sum(1 for r in c['rows'] if any(
        kw in r['desc_raw'] for kw in ['HAZ', 'DG CARGO', 'CLASS 5', 'CLASS 6', 'CLASS 8', 'CORROSIVE', 'FLAMMABLE', 'TOXIC', 'OXIDIZ']))
    hazmat_pct = hazmat_count / max(c['shipment_count'], 1)
    if hazmat_pct > 0.5:
        penalty += 20
    elif hazmat_pct > 0.1:
        penalty += 10
    return max(0, 100 - penalty)


def _classify_variance(c):
    prices = c.get('unit_prices_clean', c.get('unit_prices', []))
    if len(prices) < 5:
        return 'neutral'
    country_medians = [median(pp) for pp in c['price_by_country'].values() if len(pp) >= 2]
    month_medians = [median(pp) for pp in c['price_by_month'].values() if len(pp) >= 2]
    geo_var = _iqr(country_medians) if len(country_medians) >= 2 else 0
    time_var = _iqr(month_medians) if len(month_medians) >= 2 else 0
    overall_med = median(prices)
    if overall_med == 0:
        return 'neutral'
    geo_cv = geo_var / overall_med
    time_cv = time_var / overall_med
    if geo_cv > 0.3 and geo_cv > time_cv:
        return 'opportunity'
    if time_cv > 0.3 and time_cv > geo_cv:
        return 'risk'
    return 'neutral'


def stage4_analyse(chemicals, log=print):
    log("\nSTAGE 4 — Running 8 individual analyses...")
    chem_names = list(chemicals.keys())
    n = len(chem_names)
    if n == 0:
        return {}
    scores = {}
    for cid in chem_names:
        c = chemicals[cid]
        scores[cid] = {
            'volume': _score_volume(c), 'price': _score_price(c),
            'buyers': _score_buyers(c), 'suppliers': _score_suppliers(c),
            'trend': _score_trend(c), 'structure': _score_structure(c),
            'freedom': _score_freedom(c), 'barrier': _score_barrier(c),
            'variance_type': _classify_variance(c), 'variance_mod': 0,
        }
        vt = scores[cid]['variance_type']
        scores[cid]['variance_mod'] = 5 if vt == 'opportunity' else (-10 if vt == 'risk' else 0)
    for dim in ['volume', 'price', 'buyers', 'suppliers', 'trend', 'structure', 'freedom', 'barrier']:
        raw_vals = sorted([(cid, scores[cid][dim]) for cid in chem_names], key=lambda x: x[1])
        for rank, (cid, _) in enumerate(raw_vals):
            scores[cid][dim + '_norm'] = (rank / max(n - 1, 1)) * 100
    log(f"  Scored {n} chemicals across 8 dimensions")
    return scores


# ══════════════════════════════════════════════════════════════
# STAGE 4b — GEOPOLITICAL ADJUSTMENT
# ══════════════════════════════════════════════════════════════
def _match_geo_event(month, cid, c):
    cid_lower = cid.lower()
    countries = set(c['seller_countries'].keys()) | set(c['buyer_countries'].keys())
    for evt in GEO_EVENTS:
        if evt['start'] <= month <= evt['end']:
            country_match = bool(countries & evt['countries'])
            keyword_match = any(kw in cid_lower for kw in evt['keywords'])
            if country_match or keyword_match:
                return evt['event']
    return None


def stage4b_geo_adjust(chemicals, scores, log=print):
    log("\nSTAGE 4b — Geopolitical adjustment...")
    geo_log = []
    for cid, c in chemicals.items():
        months = sorted(m for m in c['monthly_qty'].keys() if c['monthly_qty'][m] > 0)
        if len(months) < 4:
            continue
        vals = [c['monthly_qty'][m] for m in months]
        avg = mean(vals)
        if avg == 0:
            continue
        sd = stdev(vals) if len(vals) > 1 else 0
        if sd == 0:
            continue

        anomalies = []
        for m, v in zip(months, vals):
            z = (v - avg) / sd
            if abs(z) > 2.0:
                direction = 'spike' if z > 0 else 'drop'
                event_match = _match_geo_event(m, cid, c)
                anomalies.append({
                    'month': m, 'value': v, 'z_score': z, 'direction': direction,
                    'avg': avg, 'deviation_pct': ((v - avg) / avg) * 100,
                    'event': event_match,
                })

        if anomalies:
            clean_vals = [v for m2, v in zip(months, vals)
                          if not any(a['month'] == m2 for a in anomalies)]
            if len(clean_vals) >= 3:
                adj_factor = max(0.5, min(1.5, mean(clean_vals) / avg))
            else:
                adj_factor = 1.0

            for anom in anomalies:
                geo_log.append({
                    'chemical': cid, 'month': anom['month'], 'direction': anom['direction'],
                    'z_score': round(anom['z_score'], 2),
                    'deviation_pct': round(anom['deviation_pct'], 1),
                    'raw_value': round(anom['value'], 1), 'avg_value': round(avg, 1),
                    'adj_factor': round(adj_factor, 3),
                    'event': anom['event'] or 'No matching event found — review manually',
                })

            if cid in scores:
                scores[cid]['geo_adj'] = adj_factor
                scores[cid]['trend_adjusted'] = scores[cid].get('trend_norm', 50) * adj_factor
        else:
            if cid in scores:
                scores[cid]['geo_adj'] = 1.0
                scores[cid]['trend_adjusted'] = scores[cid].get('trend_norm', 50)

    for cid in scores:
        if 'geo_adj' not in scores[cid]:
            scores[cid]['geo_adj'] = 1.0
            scores[cid]['trend_adjusted'] = scores[cid].get('trend_norm', 50)

    log(f"  Anomalies: {len(geo_log)} across {len(set(g['chemical'] for g in geo_log))} chemicals")
    return scores, geo_log


# ══════════════════════════════════════════════════════════════
# STAGE 5 — REGULATORY
# ══════════════════════════════════════════════════════════════
def stage5_regulatory(chemicals, scores, log=print):
    log("\nSTAGE 5 — Regulatory compliance check...")
    reg_log = []
    for cid in scores:
        name_l = cid.lower().strip()
        factor = REGULATORY.get(name_l, None)
        if factor is None:
            for reg_name, reg_factor in REGULATORY.items():
                if reg_name in name_l or name_l in reg_name:
                    factor = reg_factor
                    break
        if factor is None:
            factor = 1.0
        c = chemicals.get(cid, {})
        if factor >= 1.0 and c:
            hazmat = sum(1 for r in c.get('rows', []) if any(
                kw in r.get('desc_raw', '') for kw in ['TOXIC', 'CLASS 6', 'POISON', 'CARCINOGEN']))
            if hazmat > c.get('shipment_count', 1) * 0.3:
                factor = min(factor, 0.7)
        scores[cid]['reg_factor'] = factor
        status = 'clear' if factor >= 1.0 else ('conditional' if factor >= 0.7 else ('restricted' if factor > 0 else 'banned'))
        note = ''
        if status == 'restricted':
            note = 'REACH SVHC / heavy metal / carcinogen — restricted in EU, requires authorization'
        elif status == 'banned':
            note = 'Globally banned substance — do not trade'
        elif status == 'conditional':
            note = 'Requires special handling license, DG classification, or import permit'
        reg_log.append({'chemical': cid, 'factor': factor, 'status': status, 'note': note})
    restricted = sum(1 for r in reg_log if r['status'] != 'clear')
    log(f"  Clear: {len(reg_log)-restricted}, Restricted/conditional/banned: {restricted}")
    return scores, reg_log


# ══════════════════════════════════════════════════════════════
# STAGE 6 — COMPOSITE SCORING
# ══════════════════════════════════════════════════════════════
def stage6_score(chemicals, scores, log=print, weights=None, tier_a=None, tier_b=None):
    log("\nSTAGE 6 — Computing composite scores...")
    w = weights or WEIGHTS
    ta = TIER_A_MIN if tier_a is None else tier_a
    tb = TIER_B_MIN if tier_b is None else tier_b
    for cid in scores:
        s = scores[cid]
        att_base = sum(w[dim] * s.get(dim + '_norm', 0) for dim in
                       ['volume', 'price', 'buyers', 'suppliers', 'structure', 'freedom', 'barrier'])
        att_base += w['trend'] * s.get('trend_adjusted', s.get('trend_norm', 50))
        att_final = max(0, min(100, att_base * s.get('reg_factor', 1.0) + s.get('variance_mod', 0)))

        c = chemicals.get(cid, {})
        rodtep_bonus = min(5, mean(c['rodtep_values']) * 2) if c.get('rodtep_values') else 0
        drawback_bonus = min(5, mean(c['drawback_values'])) if c.get('drawback_values') else 0
        att_india = att_final + rodtep_bonus + drawback_bonus

        s.update({
            'att_base': round(att_base, 2), 'att_final': round(att_final, 2),
            'att_india': round(att_india, 2),
            'rodtep_bonus': round(rodtep_bonus, 2), 'drawback_bonus': round(drawback_bonus, 2),
            'tier': 'A' if att_final >= ta else ('B' if att_final >= tb else 'C'),
        })
    tiers = Counter(s['tier'] for s in scores.values())
    log(f"  Tier A: {tiers.get('A',0)}, Tier B: {tiers.get('B',0)}, Tier C: {tiers.get('C',0)}")
    return scores


# ══════════════════════════════════════════════════════════════
# Trend direction / growth (Time Trends logic, configurable exclusion)
# ══════════════════════════════════════════════════════════════
def compute_trend_direction(c, trend_exclude):
    """Port of the Time Trends sheet growth logic. trend_exclude is a set of
    'YYYY-MM' months excluded from the growth calculation (still displayed)."""
    months = sorted(m for m in c['monthly_shipments'] if c['monthly_shipments'][m] > 0)
    trend_vals = [c['monthly_shipments'][m] for m in months if m not in trend_exclude]
    if len(trend_vals) >= 6:
        recent = mean(trend_vals[-6:])
        prior_slice = trend_vals[-12:-6] if len(trend_vals) > 6 else trend_vals[:-6]
        prior = mean(prior_slice) if prior_slice else mean(trend_vals[:len(trend_vals) // 2])
        if prior > 0:
            growth = ((recent - prior) / prior) * 100
            direction = 'Growing' if growth > 5 else ('Declining' if growth < -5 else 'Stable')
        else:
            growth = 0
            direction = 'New'
    elif len(trend_vals) >= 3:
        half = len(trend_vals) // 2
        first = mean(trend_vals[:half])
        second = mean(trend_vals[half:])
        growth = ((second - first) / first * 100) if first > 0 else 0
        direction = 'Growing' if growth > 5 else ('Declining' if growth < -5 else 'Stable')
    else:
        direction = 'Insufficient data'
        growth = 0
    return direction, round(growth, 1)


def price_stats(c):
    """Price Deep Dive stats for one chemical."""
    clean = c.get('unit_prices_clean', [])
    raw = c.get('unit_prices', [])
    if not clean:
        return None
    med = median(clean)
    cv = (stdev(clean) / med * 100) if len(clean) > 1 and med > 0 else 0
    sp = sorted(clean)
    p5 = sp[max(0, int(len(sp) * 0.05))]
    p95 = sp[min(len(sp) - 1, int(len(sp) * 0.95))]
    country_prices = {co: median(pp) for co, pp in c.get('price_by_country', {}).items() if len(pp) >= 2}
    top_co = max(country_prices, key=country_prices.get) if country_prices else ''
    low_co = min(country_prices, key=country_prices.get) if country_prices else ''
    spread = ''
    if top_co and low_co and top_co != low_co:
        spread = f"Buy {low_co} (${country_prices[low_co]:,.1f}) -> sell at ${country_prices[top_co]:,.1f} ({top_co})"
    return {
        'raw_prices': len(raw), 'outliers_removed': c.get('outliers_removed', 0),
        'clean_prices': len(clean), 'median': round(med, 2),
        'p5': round(p5, 2), 'p95': round(p95, 2),
        'min': round(min(clean), 2), 'max': round(max(clean), 2),
        'iqr': round(_iqr(clean), 2), 'cv_pct': round(cv, 1),
        'highest_country': f"{top_co} (${country_prices.get(top_co, 0):,.1f})" if top_co else '',
        'lowest_country': f"{low_co} (${country_prices.get(low_co, 0):,.1f})" if low_co else '',
        'spread_opportunity': spread,
        'country_medians': {co: round(v, 2) for co, v in country_prices.items()},
    }


def opportunity_reasoning(c, s):
    """Opportunity Map reasoning text (ported)."""
    reasons = []
    if c['total_value_usd'] > 500000:
        reasons.append(f"High trade value (${c['total_value_usd']:,.0f})")
    elif c['total_value_usd'] > 50000:
        reasons.append(f"Moderate trade value (${c['total_value_usd']:,.0f})")
    else:
        reasons.append(f"Low trade value (${c['total_value_usd']:,.0f})")
    if len(c['buyers']) > 10:
        reasons.append(f"Strong buyer diversity ({len(c['buyers'])} buyers)")
    elif len(c['buyers']) > 3:
        reasons.append(f"Moderate buyer base ({len(c['buyers'])} buyers)")
    else:
        reasons.append(f"Narrow buyer base ({len(c['buyers'])} buyers)")
    if s.get('variance_type') == 'opportunity':
        reasons.append("Geographic price arbitrage detected")
    elif s.get('variance_type') == 'risk':
        reasons.append("Price instability risk")
    if c['shipment_count'] >= 20:
        reasons.append("High frequency trade")
    if s.get('reg_factor', 1.0) < 1.0:
        reasons.append(f"Regulatory concern (factor={s['reg_factor']:.1f})")
    return '; '.join(reasons)


def run_pipeline(exim_files, base_file, log=print, progress=None, llm_matcher=None,
                 weights=None, tier_a=None, tier_b=None):
    """Run stages 1-6. Returns a dict with all intermediate + final artifacts.
    progress(stage_name, pct) is called between stages."""
    def _p(stage, pct):
        if progress:
            progress(stage, pct)

    _p('Ingesting EXIM files', 5)
    exim_rows, base_chemicals, hsn_to_base, skipped = stage1_ingest(exim_files, base_file, log)
    _p('NLP chemical matching', 25)
    exim_rows, match_stats = stage2_nlp_match(exim_rows, base_chemicals, hsn_to_base, log, llm_matcher)
    _p('Aggregating by chemical', 45)
    base_chems, opp_chems = stage3_aggregate(exim_rows, base_chemicals, log)

    _p('Scoring base chemicals', 55)
    base_scores = stage4_analyse(base_chems, log)
    base_scores, geo_log_base = stage4b_geo_adjust(base_chems, base_scores, log)
    base_scores, reg_log_base = stage5_regulatory(base_chems, base_scores, log)
    base_scores = stage6_score(base_chems, base_scores, log, weights, tier_a, tier_b)

    _p('Scoring opportunity chemicals', 70)
    opp_scores = stage4_analyse(opp_chems, log)
    opp_scores, geo_log_opp = stage4b_geo_adjust(opp_chems, opp_scores, log)
    opp_scores, reg_log_opp = stage5_regulatory(opp_chems, opp_scores, log)
    opp_scores = stage6_score(opp_chems, opp_scores, log, weights, tier_a, tier_b)

    return {
        'exim_rows': exim_rows,
        'base_chemicals': base_chemicals,
        'base_chems': base_chems, 'base_scores': base_scores,
        'opp_chems': opp_chems, 'opp_scores': opp_scores,
        'geo_log': geo_log_base + geo_log_opp,
        'reg_log': reg_log_base + reg_log_opp,
        'match_stats': match_stats,
        'skipped_files': skipped,
    }
