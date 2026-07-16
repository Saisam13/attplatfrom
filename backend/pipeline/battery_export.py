"""Battery procurement results workbook (Suppliers, Buyers, Categories,
Monthly Trends, Raw Data)."""
from . import patches  # noqa: F401

import openpyxl

from .export import _hdr_row, _body_cell, _method_row, FILL_TIER_A, FILL_TIER_B, FILL_TIER_C

BATTERY_METHODOLOGY = {
    'Suppliers': 'METHODOLOGY: Procurement attractiveness per supplier = absolute score (not a leaderboard '
                 'percentile — a score is stable even as other suppliers are added or removed) blending '
                 'Volume 30% (log-anchored to shipment kg) + Price competitiveness 25% (median price vs category '
                 'market median — needs >=3 priced shipments in a coherent category or is excluded, not assumed '
                 '"market price") + Consistency 20% (active months / max(span, 6) — a single active month can no '
                 'longer score a perfect 1.0) + Reliability 15% (log-anchored shipment count) + Geography 10% '
                 '(trade-ease of origin country). Tier A>=70, B=40-69, C<40. Price index <1.0 means the supplier '
                 'sells below the market median for its categories; blank = insufficient price data.',
    'Buyers': 'METHODOLOGY: Competing-buyer presence = absolute score blending Volume 40% + Reliability 25% + '
              'Consistency 20% + Geography 15% (same log-anchor / floor methodology as Suppliers). These are the '
              'entities competing for the same feedstock — high-tier buyers indicate contested material.',
    'Categories': 'METHODOLOGY: Feedstock categories classified from EXIM descriptions (keyword rules: black mass, '
                  'Li-ion/lead-acid/mixed battery scrap, electrode scrap, spent catalyst, NdFeB magnet scrap, '
                  'e-waste/PCB, residues/tailings) with HSN-prefix fallback (854810, 8549, 2620/2621, 8507).',
}


def write_battery_workbook(res, output_path):
    wb = openpyxl.Workbook()

    # ── Suppliers (procurement view) ──
    ws = wb.create_sheet('Suppliers')
    cols = ['Rank', 'Supplier', 'Country', 'Categories', 'Shipments', 'Qty (KG)', 'Value (USD)',
            'Median Price', 'Price Index', 'Months Active', 'First', 'Last', 'Consistency',
            'Proc Score', 'Tier']
    _method_row(ws, 1, BATTERY_METHODOLOGY['Suppliers'], len(cols))
    _hdr_row(ws, 2, cols)
    for i, s in enumerate(res['suppliers'], 1):
        r = i + 2
        vals = [i, s['name'], s['country'], s['categories'], s['shipments'], s['qty_kg'],
                s['value_usd'], s['median_price'], s['price_index'], s['months_active'],
                s['first_month'], s['last_month'], s['consistency'], s['proc_score'], s['tier']]
        for cn, v in enumerate(vals, 1):
            _body_cell(ws, r, cn, v)
        tier_cell = ws.cell(row=r, column=len(cols))
        tier_cell.fill = {'A': FILL_TIER_A, 'B': FILL_TIER_B, 'C': FILL_TIER_C}[s['tier']]
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['D'].width = 45

    # ── Buyers (competition view) ──
    ws = wb.create_sheet('Buyers')
    _method_row(ws, 1, BATTERY_METHODOLOGY['Buyers'], len(cols))
    _hdr_row(ws, 2, cols)
    for i, s in enumerate(res['buyers'], 1):
        r = i + 2
        vals = [i, s['name'], s['country'], s['categories'], s['shipments'], s['qty_kg'],
                s['value_usd'], s['median_price'], s['price_index'], s['months_active'],
                s['first_month'], s['last_month'], s['consistency'], s['proc_score'], s['tier']]
        for cn, v in enumerate(vals, 1):
            _body_cell(ws, r, cn, v)
        tier_cell = ws.cell(row=r, column=len(cols))
        tier_cell.fill = {'A': FILL_TIER_A, 'B': FILL_TIER_B, 'C': FILL_TIER_C}[s['tier']]
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['D'].width = 45

    # ── Categories ──
    ws = wb.create_sheet('Categories')
    cat_cols = ['Category', 'Shipments', 'Qty (KG)', 'Value (USD)', 'Median Price',
                'Suppliers', 'Buyers', 'Top Origin Countries']
    _method_row(ws, 1, BATTERY_METHODOLOGY['Categories'], len(cat_cols))
    _hdr_row(ws, 2, cat_cols)
    for i, c in enumerate(res['categories'], 1):
        r = i + 2
        top = ', '.join(f'{co} ({n})' for co, n in c['top_countries'][:5])
        for cn, v in enumerate([c['category'], c['shipments'], c['qty_kg'], c['value_usd'],
                                c['median_price'], c['n_suppliers'], c['n_buyers'], top], 1):
            _body_cell(ws, r, cn, v)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['H'].width = 50

    # ── Monthly Trends per category ──
    ws = wb.create_sheet('Monthly Trends')
    months = sorted({m for c in res['categories'] for m in c['monthly_shipments']})
    _hdr_row(ws, 1, ['Category'] + months)
    for i, c in enumerate(res['categories'], 1):
        _body_cell(ws, i + 1, 1, c['category'])
        for j, m in enumerate(months, 2):
            _body_cell(ws, i + 1, j, c['monthly_shipments'].get(m, 0))
    ws.column_dimensions['A'].width = 34

    # ── Raw Data ──
    ws = wb.create_sheet('Raw Data')
    raw_cols = ['Date', 'HSN', 'Category', 'Description', 'Seller', 'Seller Country',
                'Buyer', 'Buyer Country', 'Qty', 'Qty (KG)', 'Value (USD)', 'Unit Price', 'File']
    _hdr_row(ws, 1, raw_cols)
    for i, rx in enumerate(res['rows'], 1):
        for cn, v in enumerate([rx['date'], rx['hsn6'], rx['category'], rx['desc_clean'][:200],
                                rx['seller'], rx['seller_country'], rx['buyer'], rx['buyer_country'],
                                rx['qty'], round(rx['qty_kg'], 1), round(rx['value_usd'], 1),
                                rx['unit_price'], rx['file']], 1):
            _body_cell(ws, i + 1, cn, v)
    ws.column_dimensions['D'].width = 60

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(output_path)
    return wb.sheetnames
