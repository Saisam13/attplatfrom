"""Stage 7 — 9-tab results workbook writer. Ported from trading_module.py v2.0.
Trend-exclusion months are configurable per run."""
from . import patches  # noqa: F401

from statistics import median, stdev, mean
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .constants import METHODOLOGY
from .engine import _iqr

HF = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
BODY_F = Font(name='Calibri', size=10)
METHOD_F = Font(name='Calibri', size=9, italic=True, color='555555')
FILL_HDR = PatternFill('solid', fgColor='2C3E50')
FILL_METHOD = PatternFill('solid', fgColor='F2F3F4')
FILL_TIER_A = PatternFill('solid', fgColor='27AE60')
FILL_TIER_B = PatternFill('solid', fgColor='F39C12')
FILL_TIER_C = PatternFill('solid', fgColor='BDC3C7')
FILL_OPP_HDR = PatternFill('solid', fgColor='8E44AD')
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')


def _hdr_row(ws, row, cols, fill=FILL_HDR):
    for i, col in enumerate(cols, 1):
        cl = ws.cell(row=row, column=i, value=col)
        cl.font = HF
        cl.fill = fill
        cl.alignment = CENTER
        cl.border = BORDER


def _body_cell(ws, row, col, value, fmt=None):
    cl = ws.cell(row=row, column=col, value=value)
    cl.font = BODY_F
    cl.alignment = WRAP
    cl.border = BORDER
    if fmt:
        cl.number_format = fmt
    return cl


def _method_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = METHOD_F
    cl.fill = FILL_METHOD
    cl.alignment = WRAP
    ws.row_dimensions[row].height = 40


def stage7_output(base_chems, base_scores, opp_chems, opp_scores, geo_log, reg_log,
                  exim_rows, output_path, trend_exclude=None):
    all_chems = {**base_chems, **opp_chems}
    all_scores = {**base_scores, **opp_scores}
    trend_exclude = set(trend_exclude or [])

    wb = openpyxl.Workbook()
    _write_rankings(wb, base_chems, base_scores, opp_chems, opp_scores)
    _write_price_deep(wb, all_chems, all_scores)
    _write_buyers(wb, all_chems)
    _write_suppliers(wb, all_chems)
    _write_trends(wb, all_chems, trend_exclude)
    _write_opportunity(wb, opp_chems, opp_scores)
    _write_india(wb, all_chems, all_scores)
    _write_reg_geo(wb, reg_log, geo_log)
    _write_raw(wb, exim_rows)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(output_path)
    return wb.sheetnames


def _write_rankings(wb, base_chems, base_scores, opp_chems, opp_scores):
    ws = wb.active
    ws.title = 'Rankings'
    ncols = 19
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    t = ws.cell(row=1, column=1, value='CHEMICAL TRADING ATTRACTIVENESS — RANKED OUTPUT')
    t.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor='1A5276')
    t.alignment = CENTER
    _method_row(ws, 2, METHODOLOGY['Rankings'], ncols)

    cols = ['Rank', 'Chemical', 'HSN Code(s)', 'Shipments', 'Total Qty (KG)',
            'Total Value (USD)', 'Volume', 'Price', 'Buyers', 'Suppliers',
            'Trend', 'Structure', 'Freedom', 'Barrier', 'Variance',
            'Reg Factor', 'ATT Score', 'ATT India', 'Tier']
    _hdr_row(ws, 3, cols)

    base_ranked = sorted(base_scores.items(), key=lambda x: x[1]['att_final'], reverse=True)
    r = 4
    rank = 1
    for cid, s in base_ranked:
        _write_ranking_row(ws, r, rank, cid, s, base_chems.get(cid, {}))
        r += 1
        rank += 1

    ws.merge_cells(f'A{r}:{get_column_letter(ncols)}{r}')
    sep = ws.cell(row=r, column=1, value='--- OPPORTUNITY CHEMICALS (not in base portfolio) ---')
    sep.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    sep.fill = FILL_OPP_HDR
    sep.alignment = CENTER
    r += 1

    opp_ranked = sorted(opp_scores.items(), key=lambda x: x[1]['att_final'], reverse=True)
    for cid, s in opp_ranked:
        _write_ranking_row(ws, r, rank, cid, s, opp_chems.get(cid, {}))
        r += 1
        rank += 1

    widths = [6, 40, 18, 10, 14, 14, 8, 8, 8, 8, 8, 8, 8, 8, 12, 8, 10, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}{r-1}'
    ws.freeze_panes = 'A4'


def _write_ranking_row(ws, r, rank, cid, s, c):
    _body_cell(ws, r, 1, rank)
    _body_cell(ws, r, 2, cid)
    _body_cell(ws, r, 3, ', '.join(sorted(c.get('hsn_codes', set()))))
    _body_cell(ws, r, 4, c.get('shipment_count', 0), '#,##0')
    _body_cell(ws, r, 5, round(c.get('total_qty_kg', 0)), '#,##0')
    _body_cell(ws, r, 6, round(c.get('total_value_usd', 0)), '$#,##0')
    _body_cell(ws, r, 7, round(s.get('volume_norm', 0), 1), '0.0')
    _body_cell(ws, r, 8, round(s.get('price_norm', 0), 1), '0.0')
    _body_cell(ws, r, 9, round(s.get('buyers_norm', 0), 1), '0.0')
    _body_cell(ws, r, 10, round(s.get('suppliers_norm', 0), 1), '0.0')
    _body_cell(ws, r, 11, round(s.get('trend_adjusted', 50), 1), '0.0')
    _body_cell(ws, r, 12, round(s.get('structure_norm', 0), 1), '0.0')
    _body_cell(ws, r, 13, round(s.get('freedom_norm', 0), 1), '0.0')
    _body_cell(ws, r, 14, round(s.get('barrier_norm', 0), 1), '0.0')
    _body_cell(ws, r, 15, s.get('variance_type', 'neutral'))
    _body_cell(ws, r, 16, s.get('reg_factor', 1.0), '0.0')
    cl = _body_cell(ws, r, 17, s.get('att_final', 0), '0.00')
    cl.font = Font(name='Calibri', bold=True, size=11)
    _body_cell(ws, r, 18, s.get('att_india', 0), '0.00')
    tier = s.get('tier', 'C')
    tc = _body_cell(ws, r, 19, f'Tier {tier}')
    if tier == 'A':
        tc.fill = FILL_TIER_A
        tc.font = Font(name='Calibri', bold=True, color='FFFFFF')
    elif tier == 'B':
        tc.fill = FILL_TIER_B
        tc.font = Font(name='Calibri', bold=True, color='FFFFFF')
    else:
        tc.fill = FILL_TIER_C


def _write_price_deep(wb, chemicals, scores):
    ws = wb.create_sheet('Price Deep Dive')
    cols = ['Chemical', 'Raw Prices', 'Outliers Removed', 'Clean Prices', 'Median $/unit',
            'P5 (floor)', 'P95 (cap)', 'Min (clean)', 'Max (clean)', 'IQR', 'CV %',
            'Variance Type', 'Highest Country', 'Lowest Country', 'Spread Opportunity']
    ncols = len(cols)
    _method_row(ws, 1, METHODOLOGY['Price Deep Dive'], ncols)
    _hdr_row(ws, 2, cols)
    ranked = sorted(scores.items(), key=lambda x: x[1]['att_final'], reverse=True)
    r = 3
    for cid, s in ranked:
        c = chemicals.get(cid, {})
        raw_prices = c.get('unit_prices', [])
        clean_prices = c.get('unit_prices_clean', [])
        if not clean_prices:
            continue
        med = median(clean_prices)
        mn = min(clean_prices)
        mx = max(clean_prices)
        iqr = _iqr(clean_prices)
        cv = (stdev(clean_prices) / med * 100) if len(clean_prices) > 1 and med > 0 else 0
        sp = sorted(clean_prices)
        p5 = sp[max(0, int(len(sp) * 0.05))]
        p95 = sp[min(len(sp) - 1, int(len(sp) * 0.95))]
        country_prices = {co: median(pp) for co, pp in c.get('price_by_country', {}).items() if len(pp) >= 2}
        top_co = max(country_prices, key=country_prices.get) if country_prices else ''
        low_co = min(country_prices, key=country_prices.get) if country_prices else ''
        spread = ''
        if top_co and low_co and top_co != low_co:
            spread = f"Buy {low_co} (${country_prices[low_co]:,.1f}) → sell at ${country_prices[top_co]:,.1f} ({top_co})"
        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, len(raw_prices), '#,##0')
        _body_cell(ws, r, 3, c.get('outliers_removed', 0), '#,##0')
        _body_cell(ws, r, 4, len(clean_prices), '#,##0')
        _body_cell(ws, r, 5, round(med, 2), '$#,##0.00')
        _body_cell(ws, r, 6, round(p5, 2), '$#,##0.00')
        _body_cell(ws, r, 7, round(p95, 2), '$#,##0.00')
        _body_cell(ws, r, 8, round(mn, 2), '$#,##0.00')
        _body_cell(ws, r, 9, round(mx, 2), '$#,##0.00')
        _body_cell(ws, r, 10, round(iqr, 2), '$#,##0.00')
        _body_cell(ws, r, 11, round(cv, 1), '0.0')
        _body_cell(ws, r, 12, s.get('variance_type', 'neutral'))
        _body_cell(ws, r, 13, f"{top_co} (${country_prices.get(top_co,0):,.1f})" if top_co else '')
        _body_cell(ws, r, 14, f"{low_co} (${country_prices.get(low_co,0):,.1f})" if low_co else '')
        _body_cell(ws, r, 15, spread)
        r += 1
    for i, w in enumerate([40, 8, 8, 8, 12, 10, 10, 10, 10, 10, 8, 12, 25, 25, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


def _write_buyers(wb, chemicals):
    ws = wb.create_sheet('Buyer Intel')
    cols = ['Chemical', 'Total Buyers', 'Countries', 'Top Buyer', 'Top Buyer Shipments',
            'Top Country', 'Top Country Shipments', 'Repeat Buyers %', 'Buyer List (top 10)']
    ncols = len(cols)
    _method_row(ws, 1, METHODOLOGY['Buyer Intel'], ncols)
    _hdr_row(ws, 2, cols)
    r = 3
    for cid, c in sorted(chemicals.items(), key=lambda x: len(x[1]['buyers']), reverse=True):
        tb = c['buyers'].most_common(1)[0] if c['buyers'] else ('', 0)
        tc = c['buyer_countries'].most_common(1)[0] if c['buyer_countries'] else ('', 0)
        repeat_pct = sum(1 for cnt in c['buyers'].values() if cnt > 1) / max(len(c['buyers']), 1) * 100
        top10 = '; '.join(f"{b} ({n})" for b, n in c['buyers'].most_common(10))
        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, len(c['buyers']), '#,##0')
        _body_cell(ws, r, 3, len(c['buyer_countries']), '#,##0')
        _body_cell(ws, r, 4, tb[0])
        _body_cell(ws, r, 5, tb[1], '#,##0')
        _body_cell(ws, r, 6, tc[0])
        _body_cell(ws, r, 7, tc[1], '#,##0')
        _body_cell(ws, r, 8, round(repeat_pct, 1), '0.0')
        _body_cell(ws, r, 9, top10)
        r += 1
    for i, w in enumerate([40, 10, 10, 30, 10, 15, 10, 10, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


def _write_suppliers(wb, chemicals):
    ws = wb.create_sheet('Supplier Intel')
    cols = ['Chemical', 'Total Suppliers', 'Countries', 'India Shipments', 'India %',
            'Top Supplier', 'Top Origin Country', 'Supplier List (top 10)']
    ncols = len(cols)
    _method_row(ws, 1, METHODOLOGY['Supplier Intel'], ncols)
    _hdr_row(ws, 2, cols)
    r = 3
    for cid, c in sorted(chemicals.items(), key=lambda x: len(x[1]['sellers']), reverse=True):
        ts = c['sellers'].most_common(1)[0] if c['sellers'] else ('', 0)
        to = c['seller_countries'].most_common(1)[0] if c['seller_countries'] else ('', 0)
        india = c['seller_countries'].get('INDIA', 0)
        india_pct = india / max(c['shipment_count'], 1) * 100
        top10 = '; '.join(f"{s} ({n})" for s, n in c['sellers'].most_common(10))
        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, len(c['sellers']), '#,##0')
        _body_cell(ws, r, 3, len(c['seller_countries']), '#,##0')
        _body_cell(ws, r, 4, india, '#,##0')
        _body_cell(ws, r, 5, round(india_pct, 1), '0.0')
        _body_cell(ws, r, 6, f"{ts[0]} ({ts[1]})")
        _body_cell(ws, r, 7, f"{to[0]} ({to[1]})")
        _body_cell(ws, r, 8, top10)
        r += 1
    for i, w in enumerate([40, 10, 10, 10, 8, 35, 20, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


def _write_trends(wb, chemicals, trend_exclude):
    ws = wb.create_sheet('Time Trends')
    all_months_raw = sorted(set(m for c in chemicals.values() for m in c['monthly_qty'].keys()))
    all_months = [m for m in all_months_raw
                  if any(chemicals[cid]['monthly_shipments'].get(m, 0) > 0 for cid in chemicals)]
    ncols = 2 + len(all_months) + 2
    _method_row(ws, 1, METHODOLOGY['Time Trends'], ncols)
    cols = ['Chemical', 'Total Shipments'] + all_months + ['Trend Direction', 'Growth Rate %']
    _hdr_row(ws, 2, cols)

    from statistics import mean as _mean
    r = 3
    for cid, c in sorted(chemicals.items(), key=lambda x: x[1]['shipment_count'], reverse=True):
        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, c['shipment_count'], '#,##0')
        trend_vals = []
        for i, m in enumerate(all_months):
            v = c['monthly_shipments'].get(m, 0)
            _body_cell(ws, r, 3 + i, v if v > 0 else '', '#,##0')
            if v > 0 and m not in trend_exclude:
                trend_vals.append(v)
        col_t = 3 + len(all_months)
        if len(trend_vals) >= 6:
            recent = _mean(trend_vals[-6:])
            prior_slice = trend_vals[-12:-6] if len(trend_vals) > 6 else trend_vals[:-6]
            prior = _mean(prior_slice) if prior_slice else _mean(trend_vals[:len(trend_vals) // 2])
            if prior > 0:
                growth = ((recent - prior) / prior) * 100
                direction = 'Growing' if growth > 5 else ('Declining' if growth < -5 else 'Stable')
            else:
                growth = 0
                direction = 'New'
        elif len(trend_vals) >= 3:
            half = len(trend_vals) // 2
            first = _mean(trend_vals[:half])
            second = _mean(trend_vals[half:])
            growth = ((second - first) / first * 100) if first > 0 else 0
            direction = 'Growing' if growth > 5 else ('Declining' if growth < -5 else 'Stable')
        else:
            direction = 'Insufficient data'
            growth = 0
        _body_cell(ws, r, col_t, direction)
        _body_cell(ws, r, col_t + 1, round(growth, 1), '0.0')
        r += 1
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.freeze_panes = 'C3'


def _write_opportunity(wb, opp_chems, opp_scores):
    from .engine import opportunity_reasoning
    ws = wb.create_sheet('Opportunity Map')
    ncols = 14
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    t = ws.cell(row=1, column=1, value='OPPORTUNITY CHEMICALS — Not in Base Portfolio, Fully Scored')
    t.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    t.fill = FILL_OPP_HDR
    t.alignment = CENTER
    _method_row(ws, 2, METHODOLOGY['Opportunity Map'], ncols)

    cols = ['Chemical (extracted)', 'HSN Code(s)', 'Shipments', 'Total Value (USD)',
            'Median $/unit', 'Unique Buyers', 'Unique Suppliers',
            'Top Buyer Country', 'Top Seller Country',
            'ATT Score', 'Tier', 'Trend', 'Assessment', 'Reasoning']
    _hdr_row(ws, 3, cols)

    ranked = sorted(opp_scores.items(), key=lambda x: x[1]['att_final'], reverse=True)
    r = 4
    for cid, s in ranked[:300]:
        c = opp_chems.get(cid, {})
        prices = c.get('unit_prices_clean', c.get('unit_prices', []))
        med_price = median(prices) if prices else 0
        top_buyer_co = c['buyer_countries'].most_common(1)[0][0] if c['buyer_countries'] else ''
        top_seller_co = c['seller_countries'].most_common(1)[0][0] if c['seller_countries'] else ''
        att = s.get('att_final', 0)
        tier = s.get('tier', 'C')
        assessment = 'High potential' if att >= 60 else ('Moderate' if att >= 35 else 'Low priority')

        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, ', '.join(sorted(c.get('hsn_codes', set()))))
        _body_cell(ws, r, 3, c['shipment_count'], '#,##0')
        _body_cell(ws, r, 4, round(c['total_value_usd']), '$#,##0')
        _body_cell(ws, r, 5, round(med_price, 2) if med_price else '', '$#,##0.00')
        _body_cell(ws, r, 6, len(c['buyers']), '#,##0')
        _body_cell(ws, r, 7, len(c['sellers']), '#,##0')
        _body_cell(ws, r, 8, top_buyer_co)
        _body_cell(ws, r, 9, top_seller_co)
        _body_cell(ws, r, 10, round(att, 2), '0.00')
        tc = _body_cell(ws, r, 11, f'Tier {tier}')
        if tier == 'A':
            tc.fill = FILL_TIER_A
            tc.font = Font(bold=True, color='FFFFFF')
        elif tier == 'B':
            tc.fill = FILL_TIER_B
            tc.font = Font(bold=True, color='FFFFFF')
        else:
            tc.fill = FILL_TIER_C
        trend_dir = 'Growing' if s.get('trend_adjusted', 50) > 55 else ('Declining' if s.get('trend_adjusted', 50) < 45 else 'Stable')
        _body_cell(ws, r, 12, trend_dir)
        _body_cell(ws, r, 13, assessment)
        _body_cell(ws, r, 14, opportunity_reasoning(c, s))
        r += 1

    for i, w in enumerate([40, 18, 10, 14, 12, 10, 10, 15, 15, 10, 8, 10, 14, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'


def _write_india(wb, chemicals, scores):
    ws = wb.create_sheet('India Incentives')
    cols = ['Chemical', 'ATT Final', 'RoDTEP Bonus', 'Drawback Bonus', 'ATT India',
            'India Supply %', 'Key India Routes', 'Recommendation']
    ncols = len(cols)
    _method_row(ws, 1, METHODOLOGY['India Incentives'], ncols)
    _hdr_row(ws, 2, cols)
    ranked = sorted(scores.items(), key=lambda x: x[1].get('att_india', 0), reverse=True)
    r = 3
    for cid, s in ranked:
        c = chemicals.get(cid, {})
        india_pct = c.get('seller_countries', Counter()).get('INDIA', 0) / max(c.get('shipment_count', 1), 1) * 100
        routes = []
        for country, cnt in c.get('buyer_countries', Counter()).most_common(3):
            if c.get('seller_countries', Counter()).get('INDIA', 0) > 0:
                routes.append(f"India -> {country} ({cnt})")
        rec = 'Strong India export opportunity' if india_pct > 30 and s.get('att_final', 0) >= 60 else (
              'India sourcing viable' if india_pct > 10 else 'Limited India presence')
        _body_cell(ws, r, 1, cid)
        _body_cell(ws, r, 2, s.get('att_final', 0), '0.00')
        _body_cell(ws, r, 3, s.get('rodtep_bonus', 0), '0.00')
        _body_cell(ws, r, 4, s.get('drawback_bonus', 0), '0.00')
        _body_cell(ws, r, 5, s.get('att_india', 0), '0.00')
        _body_cell(ws, r, 6, round(india_pct, 1), '0.0')
        _body_cell(ws, r, 7, '; '.join(routes) if routes else 'No India routes')
        _body_cell(ws, r, 8, rec)
        r += 1
    for i, w in enumerate([40, 10, 10, 10, 10, 10, 40, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


def _write_reg_geo(wb, reg_log, geo_log):
    ws = wb.create_sheet('Regulatory + Geo Log')
    ncols = 10
    _method_row(ws, 1, METHODOLOGY['Regulatory + Geo Log'], ncols)

    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value='REGULATORY COMPLIANCE LOG').font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(row=2, column=1).fill = PatternFill('solid', fgColor='27AE60')
    _hdr_row(ws, 3, ['Chemical', 'Status', 'Factor', 'Notes'], PatternFill('solid', fgColor='1E8449'))
    r = 4
    for entry in sorted(reg_log, key=lambda x: x['factor']):
        _body_cell(ws, r, 1, entry['chemical'])
        _body_cell(ws, r, 2, entry['status'].upper())
        _body_cell(ws, r, 3, entry['factor'], '0.0')
        _body_cell(ws, r, 4, entry.get('note', ''))
        r += 1

    r += 2
    ws.merge_cells(f'A{r}:{get_column_letter(ncols)}{r}')
    ws.cell(row=r, column=1, value='GEOPOLITICAL ADJUSTMENT LOG — Anomalies with Event Correlation').font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='8E44AD')
    r += 1
    _hdr_row(ws, r, ['Chemical', 'Month', 'Direction', 'Z-Score', 'Deviation %',
                     'Raw Value', 'Avg Value', 'Adj Factor', 'Correlated Event', ''], PatternFill('solid', fgColor='6C3483'))
    r += 1
    for entry in sorted(geo_log, key=lambda x: abs(x['z_score']), reverse=True):
        _body_cell(ws, r, 1, entry['chemical'])
        _body_cell(ws, r, 2, entry['month'])
        _body_cell(ws, r, 3, entry['direction'])
        _body_cell(ws, r, 4, entry['z_score'], '0.00')
        _body_cell(ws, r, 5, entry['deviation_pct'], '0.0')
        _body_cell(ws, r, 6, entry['raw_value'], '#,##0')
        _body_cell(ws, r, 7, entry['avg_value'], '#,##0')
        _body_cell(ws, r, 8, entry['adj_factor'], '0.000')
        _body_cell(ws, r, 9, entry.get('event', ''))
        r += 1

    ws.column_dimensions['A'].width = 40
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['I'].width = 70


def _write_raw(wb, exim_rows):
    ws = wb.create_sheet('Raw Parsed Data')
    cols = ['Date', 'HSN', 'Description (cleaned)', 'Matched Chemical', 'Match Type',
            'Match Score', 'Seller', 'Seller Country', 'Buyer', 'Buyer Country',
            'Qty', 'Qty (KG est)', 'Value (USD)', 'Unit Price']
    ncols = len(cols)
    _method_row(ws, 1, METHODOLOGY['Raw Parsed Data'], ncols)
    _hdr_row(ws, 2, cols)
    max_rows = min(len(exim_rows), 50000)
    for i in range(max_rows):
        rx = exim_rows[i]
        row = i + 3
        _body_cell(ws, row, 1, rx['date'])
        _body_cell(ws, row, 2, rx['hsn6'])
        _body_cell(ws, row, 3, rx['desc_clean'][:200])
        _body_cell(ws, row, 4, rx.get('chemical_id', ''))
        _body_cell(ws, row, 5, rx.get('match_type', ''))
        _body_cell(ws, row, 6, round(rx.get('match_score', 0), 2))
        _body_cell(ws, row, 7, rx['seller'])
        _body_cell(ws, row, 8, rx['seller_country'])
        _body_cell(ws, row, 9, rx['buyer'])
        _body_cell(ws, row, 10, rx['buyer_country'])
        _body_cell(ws, row, 11, rx['qty'], '#,##0')
        _body_cell(ws, row, 12, round(rx['qty_kg']), '#,##0')
        _body_cell(ws, row, 13, round(rx['value_usd']), '$#,##0')
        _body_cell(ws, row, 14, round(rx['unit_price'], 2) if rx['unit_price'] else '', '$#,##0.00')
    if len(exim_rows) > max_rows:
        ws.cell(row=max_rows + 3, column=1, value=f'... {len(exim_rows)-max_rows} more rows truncated')
    for i, w in enumerate([10, 8, 60, 35, 8, 8, 25, 15, 25, 15, 10, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f'A2:N{max_rows+2}'
    ws.freeze_panes = 'A3'


def write_feedback_workbook(feedback_rows, output_path):
    """Export collected trader feedback for a run as xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Feedback'
    cols = ['Chemical', 'Verdict', 'User', 'Suggested Tier', 'Expected Duration', 'Comment', 'Submitted At']
    _hdr_row(ws, 1, cols)
    for r, fb in enumerate(feedback_rows, 2):
        _body_cell(ws, r, 1, fb['chemical'])
        _body_cell(ws, r, 2, fb['verdict'])
        _body_cell(ws, r, 3, fb['user_name'])
        _body_cell(ws, r, 4, fb.get('suggested_tier') or '')
        _body_cell(ws, r, 5, fb.get('expected_duration') or '')
        _body_cell(ws, r, 6, fb.get('comment') or '')
        _body_cell(ws, r, 7, fb['created_at'])
    for i, w in enumerate([40, 12, 20, 12, 15, 70, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_path)
